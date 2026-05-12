"""HTTP endpoints for direct device-pushed ingestion + CSV/Excel import."""
import csv
import io
import uuid
from datetime import date, datetime, time, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status

from app.api.deps import AuthedDevice, DBSession, require_admin
from app.core.exceptions import http_400, http_404
from app.core.logging import get_logger
from app.models.device import Device
from app.models.reading import IngestSource, QualityFlag
from app.models.sensor import IngestionMode, Sensor
from app.models.user import User
from app.schemas.reading import (
    BatchIngestPayload,
    IngestPayload,
    IngestResponse,
)
from app.services.ingestion_service import ingest_one_reading

router = APIRouter(prefix="/ingest", tags=["ingest"])
log = get_logger(__name__)


@router.post("", response_model=IngestResponse, status_code=status.HTTP_202_ACCEPTED)
async def ingest_single(
    body: IngestPayload,
    db: DBSession,
    device: AuthedDevice,
):
    if body.device_code != device.code:
        raise HTTPException(403, "device_code mismatch with API key")

    result = await ingest_one_reading(
        db,
        device_code=body.device_code,
        sensor_code=body.sensor_code,
        ts=body.ts,
        raw_value=body.value,
        quality=body.quality,
        source=IngestSource.HTTP,
        metadata=body.metadata,
        api_key_device_id=str(device.id),
    )
    return IngestResponse(accepted=1 if result else 0, rejected=0 if result else 1)


@router.post("/batch", response_model=IngestResponse, status_code=status.HTTP_202_ACCEPTED)
async def ingest_batch(
    body: BatchIngestPayload,
    db: DBSession,
    device: AuthedDevice,
):
    if body.device_code != device.code:
        raise HTTPException(403, "device_code mismatch with API key")

    accepted = 0
    rejected = 0
    errors: list[str] = []
    for r in body.readings:
        try:
            ts = datetime.fromisoformat(str(r["ts"]))
            value = float(r["value"])
            quality = QualityFlag(r.get("quality", "good"))
            result = await ingest_one_reading(
                db,
                device_code=body.device_code,
                sensor_code=str(r["sensor_code"]),
                ts=ts,
                raw_value=value,
                quality=quality,
                source=IngestSource.HTTP,
                metadata=r.get("metadata", {}),
                api_key_device_id=str(device.id),
            )
            if result:
                accepted += 1
            else:
                rejected += 1
        except Exception as e:
            rejected += 1
            errors.append(f"{r.get('sensor_code', '?')}: {e}")
    return IngestResponse(accepted=accepted, rejected=rejected, errors=errors[:50])


@router.post("/csv", response_model=IngestResponse, status_code=status.HTTP_202_ACCEPTED)
async def ingest_csv(
    db: DBSession,
    _admin: Annotated[User, Depends(require_admin)],
    device_code: str = Query(..., min_length=1),
    file: UploadFile = File(...),
):
    """CSV format: sensor_code,ts,value[,quality]"""
    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))

    required = {"sensor_code", "ts", "value"}
    if not required.issubset({f.strip() for f in (reader.fieldnames or [])}):
        raise HTTPException(400, f"CSV must contain columns: {','.join(sorted(required))}")

    accepted = rejected = 0
    errors: list[str] = []
    for row in reader:
        try:
            ts = datetime.fromisoformat(row["ts"].strip())
            value = float(row["value"])
            quality = QualityFlag(row.get("quality", "good").strip() or "good")
            r = await ingest_one_reading(
                db,
                device_code=device_code,
                sensor_code=row["sensor_code"].strip(),
                ts=ts,
                raw_value=value,
                quality=quality,
                source=IngestSource.CSV,
            )
            if r:
                accepted += 1
            else:
                rejected += 1
        except Exception as e:
            rejected += 1
            errors.append(f"row {accepted+rejected}: {e}")
    return IngestResponse(accepted=accepted, rejected=rejected, errors=errors[:50])


def _coerce_date(v) -> date | None:
    """Best-effort parse of an excel cell into a date."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y%m%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"invalid DATE value: {v!r}")


def _coerce_time(v) -> time:
    """Best-effort parse of an excel cell into a time. Defaults to 00:00:00."""
    if v is None or v == "":
        return time(0, 0, 0)
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, (int, float)):
        # openpyxl can return a float fraction-of-day for time-only cells
        seconds = int(round(float(v) * 86400))
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        return time(h % 24, m, s)
    if isinstance(v, str):
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
            try:
                return datetime.strptime(v.strip(), fmt).time()
            except ValueError:
                continue
    raise ValueError(f"invalid TIME value: {v!r}")


@router.post(
    "/manual-excel",
    response_model=IngestResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def ingest_manual_excel(
    db: DBSession,
    _admin: Annotated[User, Depends(require_admin)],
    sensor_id: Annotated[uuid.UUID, Form(...)],
    file: UploadFile = File(...),
):
    """Upload an .xlsx file with columns DATE, TIME, M (value).

    The target sensor must have ingestion_mode = manual.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise http_400("openpyxl is not installed on the server")

    sensor = await db.get(Sensor, sensor_id)
    if not sensor:
        raise http_404("Sensor not found")
    if sensor.ingestion_mode != IngestionMode.MANUAL:
        raise http_400(
            f"Sensor {sensor.code} has ingestion_mode={sensor.ingestion_mode.value}; "
            "only manual-mode sensors accept Excel uploads"
        )

    device = await db.get(Device, sensor.device_id)
    if not device:
        raise http_404("Sensor's device not found")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise http_400(f"Could not read xlsx: {e}")

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise http_400("Empty worksheet")

    # Locate columns by header name (case-insensitive). DATE/TIME/M.
    norm = [str(c).strip().upper() if c is not None else "" for c in header]
    try:
        idx_date = norm.index("DATE")
        idx_time = norm.index("TIME")
        idx_m = norm.index("M")
    except ValueError:
        raise http_400(
            f"Excel must contain columns DATE, TIME, M (got headers: {header})"
        )

    accepted = rejected = 0
    errors: list[str] = []
    row_num = 1  # header is row 1
    for row in rows:
        row_num += 1
        try:
            if row is None or all(c is None for c in row):
                continue
            d = _coerce_date(row[idx_date])
            if d is None:
                continue
            t = _coerce_time(row[idx_time])
            v_raw = row[idx_m]
            if v_raw is None or v_raw == "":
                raise ValueError("missing M value")
            value = float(v_raw)
            ts = datetime.combine(d, t).replace(tzinfo=timezone.utc)
            result = await ingest_one_reading(
                db,
                device_code=device.code,
                sensor_code=sensor.code,
                ts=ts,
                raw_value=value,
                quality=QualityFlag.GOOD,
                source=IngestSource.CSV,
                metadata={"manual_upload": True, "row": row_num},
            )
            if result:
                accepted += 1
            else:
                rejected += 1
        except Exception as e:
            rejected += 1
            errors.append(f"row {row_num}: {e}")

    log.info(
        "ingest.manual_excel",
        sensor=sensor.code,
        accepted=accepted,
        rejected=rejected,
    )
    return IngestResponse(accepted=accepted, rejected=rejected, errors=errors[:50])
